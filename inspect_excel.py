import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("PFB2073_WellTest_Results.xlsx", data_only=False)
print("=== EXCEL WORKBOOK DETAILED ANALYSIS ===")
print("Sheet Names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    print(f"Dimensions: {ws.dimensions}")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            print(f"  Row {r}: {[str(v) for v in row_vals if v is not None]}")

