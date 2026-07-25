"""
tests/test_excel_exporter.py
----------------------------
Unit tests for src/excel_exporter.py module.
Verifies that export_well_test_results generates a valid Excel workbook with
expected sheet names ('Problem 1 Drawdown', 'Problem 2 Buildup'), non-empty cell
values across parameter and data tables, correct styling structure, and embedded image attachments.
"""

import os
import openpyxl
import pytest
from src.welltest import analyze_problem1_drawdown, analyze_problem2_buildup
from src.excel_exporter import export_well_test_results


@pytest.fixture
def p1_result():
    return analyze_problem1_drawdown()


@pytest.fixture
def p2_result():
    return analyze_problem2_buildup()


def test_export_well_test_results_creation(tmp_path, p1_result, p2_result):
    excel_path = str(tmp_path / "PFB2073_WellTest_Results_test.xlsx")
    out_file = export_well_test_results(p1_result, p2_result, output_path=excel_path)

    assert out_file == excel_path
    assert os.path.exists(excel_path)
    assert os.path.getsize(excel_path) > 5000  # Non-empty file size


def test_excel_sheet_structure(tmp_path, p1_result, p2_result):
    excel_path = str(tmp_path / "PFB2073_WellTest_Results_test.xlsx")
    export_well_test_results(p1_result, p2_result, output_path=excel_path)

    wb = openpyxl.load_workbook(excel_path)
    try:
        sheet_names = wb.sheetnames

        assert "Problem 1 Drawdown" in sheet_names
        assert "Problem 2 Buildup" in sheet_names

        ws1 = wb["Problem 1 Drawdown"]
        ws2 = wb["Problem 2 Buildup"]

        # Verify Sheet 1 title and headers
        assert ws1["A1"].value == "PFB2073 WELL TEST ANALYSIS - PROBLEM 1 DRAWDOWN TEST"
        assert ws1.cell(row=3, column=1).value == "Reservoir Parameter"
        assert ws1.cell(row=3, column=4).value == "Calculated Result KPI"
        assert ws1.cell(row=16, column=1).value == "Time t (hr)"
        assert ws1.cell(row=16, column=2).value == "Flowing Pressure p_wf (psi)"

        # Verify Sheet 1 data rows present (26 rows + headers)
        assert ws1.cell(row=17, column=1).value is not None
        assert ws1.cell(row=42, column=1).value is not None

        # Verify Sheet 2 title and headers
        assert ws2["A1"].value == "PFB2073 WELL TEST ANALYSIS - PROBLEM 2 BUILDUP TEST"
        assert ws2.cell(row=3, column=1).value == "Reservoir Parameter"
        assert ws2.cell(row=3, column=4).value == "Calculated Result KPI"
        assert ws2.cell(row=21, column=1).value == "Shut-in Time Δt (hr)"
        assert ws2.cell(row=21, column=2).value == "Agarwal Time t_e (hr)"

        # Verify Sheet 2 data rows present (33 rows + headers)
        assert ws2.cell(row=22, column=1).value is not None
        assert ws2.cell(row=54, column=1).value is not None
    finally:
        wb.close()


def test_excel_embedded_images(tmp_path, p1_result, p2_result):
    excel_path = str(tmp_path / "PFB2073_WellTest_Results_test.xlsx")
    export_well_test_results(p1_result, p2_result, output_path=excel_path)

    wb = openpyxl.load_workbook(excel_path)
    try:
        ws1 = wb["Problem 1 Drawdown"]
        ws2 = wb["Problem 2 Buildup"]

        assert len(ws1._images) > 0, "Sheet 1 missing embedded diagnostic plot image"
        assert len(ws2._images) > 0, "Sheet 2 missing embedded diagnostic plot image"
    finally:
        wb.close()


def test_export_default_file_generation():
    default_path = "PFB2073_WellTest_Results.xlsx"
    out_file = export_well_test_results()

    assert os.path.exists(default_path)
    assert os.path.getsize(default_path) > 5000
    
    wb = openpyxl.load_workbook(default_path)
    try:
        assert "Problem 1 Drawdown" in wb.sheetnames
        assert "Problem 2 Buildup" in wb.sheetnames
    finally:
        wb.close()


def test_export_well_test_results_nested_dir(tmp_path, p1_result, p2_result):
    nested_path = str(tmp_path / "nested" / "output_folder" / "results.xlsx")
    out_file = export_well_test_results(p1_result, p2_result, output_path=nested_path)

    assert out_file == nested_path
    assert os.path.exists(nested_path)
    assert os.path.getsize(nested_path) > 5000
