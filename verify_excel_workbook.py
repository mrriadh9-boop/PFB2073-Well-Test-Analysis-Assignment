import openpyxl

def audit_excel():
    wb = openpyxl.load_workbook("PFB2073_WellTest_Results.xlsx", data_only=True)
    print("=== EXCEL WORKBOOK AUDIT ===")
    print("Sheet names:", wb.sheetnames)
    
    expected_sheets = ["Problem 1 Drawdown", "Problem 2 Buildup"]
    for sname in expected_sheets:
        if sname not in wb.sheetnames:
            print(f"ERROR: Expected sheet '{sname}' missing!")
            continue
        ws = wb[sname]
        print(f"\nSheet '{sname}': max_row={ws.max_row}, max_col={ws.max_column}")
        
        # Check images
        images = getattr(ws, '_images', [])
        print(f"  Embedded images count: {len(images)}")
        
        # Check top header cell
        cell_a1 = ws['A1'].value
        print(f"  Title cell A1: '{cell_a1}'")
        
        # Check calculated parameters block
        found_k = False
        found_s = False
        found_C = False
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value)
                if 'Permeability' in val or 'k (mD)' in val or 'k =' in val:
                    found_k = True
                if 'Skin' in val or 's =' in val:
                    found_s = True
                if 'Wellbore Storage' in val or 'C =' in val:
                    found_C = True
        print(f"  Contains k: {found_k}, s: {found_s}, C: {found_C}")

if __name__ == '__main__':
    audit_excel()
