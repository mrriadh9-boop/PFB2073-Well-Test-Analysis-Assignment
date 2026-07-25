"""
src/excel_exporter.py
---------------------
Excel workbook generation engine for Petroleum Engineering Well Test Analysis.
Exports complete analysis results for Drawdown (Problem 1) and Buildup (Problem 2)
to 'PFB2073_WellTest_Results.xlsx' with openpyxl formatting, styling, summary KPIs,
full data tables, and embedded high-resolution diagnostic plots.
"""

import io
import os
from typing import Optional
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from src.welltest import (
    Problem1AnalysisResult,
    Problem2AnalysisResult,
    analyze_problem1_drawdown,
    analyze_problem2_buildup,
)
from src.plotting import generate_all_plots, plot_drawdown_diagnostic, plot_buildup_diagnostic


def _apply_table_formatting(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    """Utility to apply thin borders and aligned styles to a table block."""
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")


def export_well_test_results(
    p1_res: Optional[Problem1AnalysisResult] = None,
    p2_res: Optional[Problem2AnalysisResult] = None,
    output_path: str = "PFB2073_WellTest_Results.xlsx"
) -> str:
    """
    Exports well test analysis results to an Excel file with formatted sheets,
    summary KPI blocks, parameter tables, data tables, and embedded diagnostic images.

    Parameters:
        p1_res      : Problem1AnalysisResult object (optional).
        p2_res      : Problem2AnalysisResult object (optional).
        output_path : Path where Excel workbook will be saved.

    Returns:
        str: Path to saved Excel file.
    """
    if p1_res is None:
        p1_res = analyze_problem1_drawdown()
    if p2_res is None:
        p2_res = analyze_problem2_buildup()

    # Ensure diagnostic plots are generated for embedding
    p1_plot_path = "Problem1_LogLog_Diagnostic.png"
    p2_plot_path = "Problem2_LogLog_Diagnostic.png"
    if (
        not os.path.exists(p1_plot_path) or os.path.getsize(p1_plot_path) == 0
        or not os.path.exists(p2_plot_path) or os.path.getsize(p2_plot_path) == 0
    ):
        generate_all_plots(p1_res, p2_res)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    section_font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    normal_font = Font(name="Calibri", size=10)

    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    table_hdr_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # SHEET 1: Problem 1 Drawdown
    # ==========================================
    ws1 = wb.create_sheet(title="Problem 1 Drawdown")
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "PFB2073 WELL TEST ANALYSIS - PROBLEM 1 DRAWDOWN TEST"
    ws1["A1"].font = title_font

    # Reservoir Parameters Block (Cols A-B, Rows 3-11)
    ws1.cell(row=3, column=1, value="Reservoir Parameter").fill = hdr_fill
    ws1.cell(row=3, column=1).font = hdr_font
    ws1.cell(row=3, column=1).alignment = align_center

    ws1.cell(row=3, column=2, value="Value").fill = hdr_fill
    ws1.cell(row=3, column=2).font = hdr_font
    ws1.cell(row=3, column=2).alignment = align_center

    p1_params_data = [
        ("Formation Thickness h (ft)", p1_res.params.h, "0.0"),
        ("Porosity phi (fraction)", p1_res.params.phi, "0.000"),
        ("Viscosity mu (cP)", p1_res.params.mu, "0.00"),
        ("Total Compressibility c_t (psi^-1)", p1_res.params.c_t, "0.00E+00"),
        ("Wellbore Radius r_w (ft)", p1_res.params.r_w, "0.00"),
        ("Oil Formation Volume Factor B_o (RB/STB)", p1_res.params.B, "0.000"),
        ("Oil Flow Rate q (STB/D)", p1_res.params.q, "0.0"),
        ("Initial Pressure p_i (psia)", p1_res.params.p_i, "0.00"),
    ]

    for idx, (lbl, val, fmt) in enumerate(p1_params_data, start=4):
        c_lbl = ws1.cell(row=idx, column=1, value=lbl)
        c_lbl.font = bold_font
        c_lbl.alignment = align_left

        c_val = ws1.cell(row=idx, column=2, value=val)
        c_val.font = normal_font
        c_val.number_format = fmt
        c_val.alignment = align_right

    _apply_table_formatting(ws1, 3, 11, 1, 2)

    # Calculated KPI Summary Block (Cols D-E, Rows 3-13)
    ws1.cell(row=3, column=4, value="Calculated Result KPI").fill = hdr_fill
    ws1.cell(row=3, column=4).font = hdr_font
    ws1.cell(row=3, column=4).alignment = align_center

    ws1.cell(row=3, column=5, value="Value").fill = hdr_fill
    ws1.cell(row=3, column=5).font = hdr_font
    ws1.cell(row=3, column=5).alignment = align_center

    p1_kpi_data = [
        ("Permeability k (md)", p1_res.type_curve.k, "0.00"),
        ("Skin Factor s", p1_res.type_curve.s, "0.00"),
        ("Wellbore Storage C (bbl/psi)", p1_res.type_curve.C, "0.0000"),
        ("Dimensionless Storage C_D", p1_res.type_curve.C_D, "0.00"),
        ("Gringarten Group C_D e^(2s)", p1_res.type_curve.C_D_e2s, "0.00"),
        ("Semi-Log Slope m (psi/cycle)", p1_res.semilog.m, "0.00"),
        ("Extrapolated Pressure p_1hr (psia)", p1_res.semilog.p_1hr, "0.00"),
        ("Semi-Log Permeability k_semilog (md)", p1_res.semilog.k, "0.00"),
        ("Semi-Log Skin Factor s_semilog", p1_res.semilog.s, "0.00"),
        ("Overall Fit R^2", p1_res.type_curve.r2_overall, "0.0000"),
    ]

    for idx, (lbl, val, fmt) in enumerate(p1_kpi_data, start=4):
        c_lbl = ws1.cell(row=idx, column=4, value=lbl)
        c_lbl.font = bold_font
        c_lbl.alignment = align_left

        c_val = ws1.cell(row=idx, column=5, value=val)
        c_val.font = normal_font
        c_val.number_format = fmt
        c_val.alignment = align_right

    _apply_table_formatting(ws1, 3, 13, 4, 5)

    # Data Table Section Header
    ws1.cell(row=15, column=1, value="Problem 1 Observation Data & Matched Model").font = section_font

    # Data Table Headers (Cols A-F, Row 16)
    p1_headers = [
        "Time t (hr)",
        "Flowing Pressure p_wf (psi)",
        "Pressure Drop Δp (psi)",
        "Bourdet Derivative (psi)",
        "Matched Model Δp (psi)",
        "Matched Model Derivative (psi)",
    ]

    for col_idx, h_text in enumerate(p1_headers, start=1):
        cell = ws1.cell(row=16, column=col_idx, value=h_text)
        cell.fill = table_hdr_fill
        cell.font = hdr_font
        cell.alignment = align_center

    # Populate Data Rows
    df1 = p1_res.df
    tc1 = p1_res.type_curve

    for r_idx in range(len(df1)):
        row_num = 17 + r_idx
        ws1.cell(row=row_num, column=1, value=df1['time_hr'].iloc[r_idx]).number_format = "0.000"
        ws1.cell(row=row_num, column=2, value=df1['p_wf_psi'].iloc[r_idx]).number_format = "0.00"
        ws1.cell(row=row_num, column=3, value=df1['delta_p_psi'].iloc[r_idx]).number_format = "0.00"
        ws1.cell(row=row_num, column=4, value=df1['bourdet_derivative'].iloc[r_idx]).number_format = "0.00"
        ws1.cell(row=row_num, column=5, value=tc1.dp_pred[r_idx]).number_format = "0.00"
        ws1.cell(row=row_num, column=6, value=tc1.dp_prime_pred[r_idx]).number_format = "0.00"

        for c_idx in range(1, 7):
            cell = ws1.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.alignment = align_right

    _apply_table_formatting(ws1, 16, 16 + len(df1), 1, 6)

    # Embed Diagnostic Image on Sheet 1
    if os.path.exists(p1_plot_path) and os.path.getsize(p1_plot_path) > 0:
        with open(p1_plot_path, "rb") as f1:
            img1 = Image(io.BytesIO(f1.read()))
        img1.width = 540
        img1.height = 415
        ws1.add_image(img1, "H3")

    # ==========================================
    # SHEET 2: Problem 2 Buildup
    # ==========================================
    ws2 = wb.create_sheet(title="Problem 2 Buildup")
    ws2.views.sheetView[0].showGridLines = True

    # Title
    ws2["A1"] = "PFB2073 WELL TEST ANALYSIS - PROBLEM 2 BUILDUP TEST"
    ws2["A1"].font = title_font

    # Reservoir Parameters Block (Cols A-B, Rows 3-12)
    ws2.cell(row=3, column=1, value="Reservoir Parameter").fill = hdr_fill
    ws2.cell(row=3, column=1).font = hdr_font
    ws2.cell(row=3, column=1).alignment = align_center

    ws2.cell(row=3, column=2, value="Value").fill = hdr_fill
    ws2.cell(row=3, column=2).font = hdr_font
    ws2.cell(row=3, column=2).alignment = align_center

    p2_params_data = [
        ("Formation Thickness h (ft)", p2_res.params.h, "0.0"),
        ("Porosity phi (fraction)", p2_res.params.phi, "0.000"),
        ("Viscosity mu (cP)", p2_res.params.mu, "0.00"),
        ("Total Compressibility c_t (psi^-1)", p2_res.params.c_t, "0.00E+00"),
        ("Wellbore Radius r_w (ft)", p2_res.params.r_w, "0.00"),
        ("Oil Formation Volume Factor B_o (RB/STB)", p2_res.params.B, "0.000"),
        ("Oil Flow Rate q (STB/D)", p2_res.params.q, "0.0"),
        ("Production Time t_p (hr)", p2_res.params.t_p, "0.0"),
        ("Drainage Area A (acres)", p2_res.params.A, "0.0"),
        ("Initial Flowing Pressure p_wf(0) (psia)", p2_res.params.p_i, "0.00"),
    ]

    for idx, (lbl, val, fmt) in enumerate(p2_params_data, start=4):
        c_lbl = ws2.cell(row=idx, column=1, value=lbl)
        c_lbl.font = bold_font
        c_lbl.alignment = align_left

        c_val = ws2.cell(row=idx, column=2, value=val)
        c_val.font = normal_font
        c_val.number_format = fmt
        c_val.alignment = align_right

    _apply_table_formatting(ws2, 3, 12, 1, 2)

    # Calculated KPI Summary Block (Cols D-E, Rows 3-18)
    ws2.cell(row=3, column=4, value="Calculated Result KPI").fill = hdr_fill
    ws2.cell(row=3, column=4).font = hdr_font
    ws2.cell(row=3, column=4).alignment = align_center

    ws2.cell(row=3, column=5, value="Value").fill = hdr_fill
    ws2.cell(row=3, column=5).font = hdr_font
    ws2.cell(row=3, column=5).alignment = align_center

    p2_kpi_data = [
        ("Permeability k (md)", p2_res.type_curve.k, "0.00"),
        ("Skin Factor s", p2_res.type_curve.s, "0.00"),
        ("Wellbore Storage C (bbl/psi)", p2_res.type_curve.C, "0.0000"),
        ("Dimensionless Storage C_D", p2_res.type_curve.C_D, "0.00"),
        ("Gringarten Parameter C_D e^(2s)", p2_res.type_curve.C_D_e2s, "0.00"),
        ("Horner Slope m (psi/cycle)", p2_res.semilog.m, "0.00"),
        ("Extrapolated Pressure p_1hr (psia)", p2_res.semilog.p_1hr, "0.00"),
        ("False Pressure p* (psia)", p2_res.semilog.p_star, "0.00"),
        ("True Avg Pressure p_bar (MBH/Dietz) (psia)", p2_res.semilog.p_bar, "0.00"),
        ("Semi-Log Permeability k_semilog (md)", p2_res.semilog.k, "0.00"),
        ("Semi-Log Skin Factor s_semilog", p2_res.semilog.s, "0.00"),
        ("Radius of Inv r_inv at 72h (ft)", p2_res.semilog.r_inv_72hr, "0.00"),
        ("Nearest Boundary Distance (ft)", p2_res.semilog.nearest_boundary_ft, "0.00"),
        ("Time to Reach Boundary (hr)", p2_res.semilog.time_to_boundary_hr, "0.00"),
        ("Boundary Reached at 72h", "Yes" if p2_res.semilog.boundary_reached else "No", "@"),
        ("Overall Fit R^2", p2_res.type_curve.r2_overall, "0.0000"),
    ]

    for idx, (lbl, val, fmt) in enumerate(p2_kpi_data, start=4):
        c_lbl = ws2.cell(row=idx, column=4, value=lbl)
        c_lbl.font = bold_font
        c_lbl.alignment = align_left

        c_val = ws2.cell(row=idx, column=5, value=val)
        c_val.font = normal_font
        if fmt != "@":
            c_val.number_format = fmt
        c_val.alignment = align_right

    _apply_table_formatting(ws2, 3, 19, 4, 5)

    # Data Table Section Header
    ws2.cell(row=20, column=1, value="Problem 2 Observation Data & Matched Model").font = section_font

    # Data Table Headers (Cols A-H, Row 21)
    p2_headers = [
        "Shut-in Time Δt (hr)",
        "Agarwal Time t_e (hr)",
        "Horner Time Ratio",
        "Shut-in Pressure p_ws (psi)",
        "Pressure Change Δp (psi)",
        "Bourdet Derivative (psi)",
        "Matched Model Δp (psi)",
        "Matched Model Derivative (psi)",
    ]

    for col_idx, h_text in enumerate(p2_headers, start=1):
        cell = ws2.cell(row=21, column=col_idx, value=h_text)
        cell.fill = table_hdr_fill
        cell.font = hdr_font
        cell.alignment = align_center

    # Populate Data Rows
    df2 = p2_res.df
    tc2 = p2_res.type_curve

    for r_idx in range(len(df2)):
        row_num = 22 + r_idx
        dt_val = df2['delta_t_hr'].iloc[r_idx]
        te_val = df2['t_e_hr'].iloc[r_idx]
        hr_val = df2['horner_ratio'].iloc[r_idx]
        pws_val = df2['p_ws_psi'].iloc[r_idx]
        dp_val = df2['delta_p_psi'].iloc[r_idx]
        dpp_val = df2['bourdet_derivative'].iloc[r_idx]

        if r_idx == 0:
            dp_pred_val = 0.0
            dpp_pred_val = 0.0
        else:
            dp_pred_val = tc2.dp_pred[r_idx - 1]
            dpp_pred_val = tc2.dp_prime_pred[r_idx - 1]

        ws2.cell(row=row_num, column=1, value=dt_val).number_format = "0.000"
        ws2.cell(row=row_num, column=2, value=te_val).number_format = "0.000"

        c_hr = ws2.cell(row=row_num, column=3)
        if np.isinf(hr_val):
            c_hr.value = "N/A"
        else:
            c_hr.value = hr_val
            c_hr.number_format = "0.00"

        ws2.cell(row=row_num, column=4, value=pws_val).number_format = "0.00"
        ws2.cell(row=row_num, column=5, value=dp_val).number_format = "0.00"
        ws2.cell(row=row_num, column=6, value=dpp_val).number_format = "0.00"
        ws2.cell(row=row_num, column=7, value=dp_pred_val).number_format = "0.00"
        ws2.cell(row=row_num, column=8, value=dpp_pred_val).number_format = "0.00"

        for c_idx in range(1, 9):
            cell = ws2.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.alignment = align_right

    _apply_table_formatting(ws2, 21, 21 + len(df2), 1, 8)

    # Embed Diagnostic Image on Sheet 2
    if os.path.exists(p2_plot_path) and os.path.getsize(p2_plot_path) > 0:
        with open(p2_plot_path, "rb") as f2:
            img2 = Image(io.BytesIO(f2.read()))
        img2.width = 540
        img2.height = 415
        ws2.add_image(img2, "J3")

    # ==========================================
    # Auto-adjust column widths for all sheets
    # ==========================================
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Avoid title merged wide cells skewing width
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    dirname = os.path.dirname(output_path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    wb.save(output_path)
    return output_path
